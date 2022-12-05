import csv
import json
import os
from typing import Type

import openpyxl
from openpyxl import Workbook

from data.project.base import Entity, Dataset


class CSVHandler:
    """
    A class that handles CSV documents.
    """

    @staticmethod
    def read_entity(entity_type: Type[Entity], path: str, file_name: str = None,
                    extension: str = ".csv", delimiter: str = ";") -> list[Entity]:
        """
        Reads entries from a CSV document.

        :param entity_type: the type of entries
        :param path: the path of the document
        :param file_name: the name of the document
        :param extension: the extension of the document
        :param delimiter: the delimiter
        :return: the list of elements
        """
        file_name = file_name if file_name is not None else entity_type.collection_name()

        with open(os.path.join(path, file_name + extension), encoding="utf-8") as file:
            rows = csv.DictReader(file, delimiter=delimiter)
            return [entity_type.from_sequence([row[n] for n in entity_type.field_names()]) for row in rows]

    @staticmethod
    def write_entity(entities: list[Entity], path: str, file_name: str = None,
                     extension: str = ".csv", delimiter: str = ";") -> None:
        """
        Writes entries to a CSV document.

        :param entities: the entries
        :param path: the path of the document
        :param file_name: the name of the document
        :param extension: the extension of the document
        :param delimiter: the delimiter
        :return: nothing
        """
        file_name = file_name if file_name is not None else entities[0].collection_name()


        with open(os.path.join(path, file_name + extension), "w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=entities[0].field_names(), delimiter=delimiter)
            writer.writeheader()
            for entity in entities:
                writer.writerow(entity.__dict__)

    @staticmethod
    def read_dataset(dataset_type: Type[Dataset], path: str) -> Dataset:
        """
        Reads a dataset from multiple CSV documents.

        :param dataset_type: the type of the dataset
        :param path: the path of the documents
        :return: the instance
        """
        return dataset_type.from_sequence(
            [
                CSVHandler.read_entity(entity_type, path, file_name=entity_type.collection_name())
                for entity_type in dataset_type.entity_types()
            ]
        )

    @staticmethod
    def write_dataset(dataset: Dataset, path: str) -> None:
        """
        Writes a dataset to multiple CSV documents.

        :param dataset: the dataset instance
        :param path: the path of the documents
        :return: nothing
        """
        for entity_type in dataset.entity_types():
            CSVHandler.write_entity(dataset.entities()[entity_type], path, file_name=entity_type.collection_name())


class JSONHandler:
    """
    A class that handles JSON documents.
    """

    @staticmethod
    def read_entity(entity_type: Type[Entity], path: str, file_name: str = None, extension: str = ".json") -> list[
        Entity]:
        """
        Reads entries from a JSON document.

        :param entity_type: the type of entries
        :param path: the path of the document
        :param file_name: the name of the document
        :param extension: the extension of the document
        :return: the list of elements
        """

        file_name = file_name if file_name is not None else entity_type.collection_name()

        with open(os.path.join(path, file_name + extension), encoding="utf-8") as file:
            return [entity_type.from_sequence([raw_entity[name] for name in entity_type.field_names()]) for
                    raw_entity in json.load(file)]

    @staticmethod
    def write_entity(entities: list[Entity], path: str, file_name: str = None, extension: str = ".json",
                     pretty: bool = True) -> None:
        """
        Writes entries to a CSV document.

        :param entities: the entries
        :param path: the path of the document
        :param file_name: the name of the document
        :param extension: the extension of the document
        :param pretty: tells whether the file should be indented or not
        :return: nothing
        """

        file_name = file_name if file_name is not None else entities[0].collection_name()

        with open(os.path.join(path, file_name + extension), "w", newline="", encoding="utf-8") as file:
            json.dump([entity.__dict__ for entity in entities], file, indent=2 if pretty else 0)

    @staticmethod
    def read_dataset(dataset_type: Type[Dataset], path: str) -> Dataset:
        """
        Reads a dataset from multiple JSON documents.

        :param dataset_type: the type of the dataset
        :param path: the path of the documents
        :return: the instance
        """
        return dataset_type.from_sequence(
            [
                JSONHandler.read_entity(entity_type, path, file_name=entity_type.collection_name())
                for entity_type in dataset_type.entity_types()
            ]
        )

    @staticmethod
    def write_dataset(dataset: Dataset, path: str) -> None:
        """
        Writes a dataset to multiple JSON documents.

        :param dataset: the dataset instance
        :param path: the path of the documents
        :return: nothing
        """
        for entity_type in dataset.entity_types():
            JSONHandler.write_entity(dataset.entities()[entity_type], path, file_name=entity_type.collection_name())


class XLSXHandler:
    """
    A class that handles XLSX documents.
    """

    @staticmethod
    def read_entity(entity_type: Type[Entity], workbook: openpyxl.Workbook, sheet_name: str = None,
                    heading: bool = True) -> list[Entity]:
        """
        Reads entries from an XLSX document.

        :param entity_type: the type of entries
        :param workbook: the workbook instance
        :param sheet_name: the name of the worksheet
        :param heading: tells whether a heading should be added to the worksheet
        :return: the list of elements
        """

        sheet_name = sheet_name if sheet_name is not None else entity_type.collection_name()

        sheet = workbook[sheet_name]
        entities = []

        row = 2 if heading else 1
        while True:
            cell = sheet.cell(row=row, column=1)
            if cell.value is None:
                break

            values = [sheet.cell(row=row, column=pos).value for pos in range(1, len(entity_type.field_names()) + 1)]
            entities.append(entity_type.from_sequence(values))
            row += 1

        return entities

    @staticmethod
    def write_entity(entities: list[Entity], workbook: openpyxl.Workbook, sheet_name: str = None,
                     heading: bool = True) -> None:
        """
        Writes entries to an XLSX document.

        :param entities: the entries
        :param workbook: the workbook instance
        :param sheet_name: the name of the worksheet
        :param heading: tells whether a heading can be found in the worksheet
        :return: nothing
        """

        sheet_name = sheet_name if sheet_name is not None else entities[0].collection_name()

        sheet = workbook.create_sheet(sheet_name)
        if heading:
            for i in range(len(entities[0].field_names())):
                sheet.cell(row=1, column=i + 1, value=entities[0].field_names()[i])

        row = 2 if heading else 1
        for entity in entities:
            for j in range(len(entities[0].field_names())):
                sheet.cell(row=row, column=j + 1, value=entity.__dict__[entities[0].field_names()[j]])
            row += 1

    @staticmethod
    def read_dataset(dataset_type: Type[Dataset], path: str) -> Dataset:
        """
        Reads a dataset from an XLSX document.

        :param dataset_type: the type of the dataset
        :param path: the path of the document
        :return: the instance
        """

        wb = openpyxl.load_workbook(os.path.join(path, "dataset.xlsx"))
        return dataset_type.from_sequence(
            [
                XLSXHandler.read_entity(entity_type, wb, sheet_name=entity_type.collection_name())
                for entity_type in dataset_type.entity_types()
            ]
        )

    @staticmethod
    def write_dataset(dataset: Dataset, path: str) -> None:
        """
        Writes a dataset to to an XLSX document.

        :param dataset: the dataset instance
        :param path: the path of the document
        :return: nothing
        """

        wb = Workbook()
        for entity_type in dataset.entity_types():
            XLSXHandler.write_entity(dataset.entities()[entity_type], wb, sheet_name=entity_type.collection_name())
        wb.remove(wb["Sheet"])
        wb.save(os.path.join(path, "dataset.xlsx"))

